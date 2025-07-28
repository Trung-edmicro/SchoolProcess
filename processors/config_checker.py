"""
Config Checker
Kiểm tra và validate cấu hình hệ thống
Author: Assistant
Date: 2025-07-26
"""

import os
import pandas as pd
from typing import Dict, List, Tuple
import json


class ConfigChecker:
    """Class kiểm tra cấu hình hệ thống"""
    
    def __init__(self, input_folder: str, temp_folder: str, config_folder: str = "config"):
        """
        Khởi tạo ConfigChecker
        
        Args:
            input_folder (str): Thư mục chứa file input
            temp_folder (str): Thư mục chứa template
            config_folder (str): Thư mục config
        """
        self.input_folder = input_folder
        self.temp_folder = temp_folder
        self.config_folder = config_folder
        
        # Đường dẫn các file
        self.student_file = os.path.join(input_folder, "Danh sach hoc sinh.xlsx")
        self.teacher_file = os.path.join(input_folder, "DS tài khoản giáo viên.xlsx")
        self.template_file = os.path.join(temp_folder, "Template_Export.xlsx")
        self.service_account_file = os.path.join(config_folder, "service_account.json")
        self.config_file = os.path.join(config_folder, "config.py")
        self.google_api_file = os.path.join(config_folder, "google_api.py")
    
    def check_file_system(self) -> Dict[str, Dict]:
        """
        Kiểm tra hệ thống file và thư mục
        
        Returns:
            Dict: Kết quả kiểm tra file system
        """
        print("\n📁 KIỂM TRA HỆ THỐNG FILE:")
        print("=" * 40)
        
        results = {
            "directories": {},
            "input_files": {},
            "config_files": {},
            "template_files": {}
        }
        
        # Kiểm tra thư mục
        directories = {
            "Input Folder": self.input_folder,
            "Template Folder": self.temp_folder,
            "Config Folder": self.config_folder
        }
        
        for name, path in directories.items():
            exists = os.path.exists(path)
            results["directories"][name] = {"path": path, "exists": exists}
            status = "✅" if exists else "❌"
            print(f"   {name}: {status} {path}")
        
        # Kiểm tra file input
        input_files = {
            "File học sinh": self.student_file,
            "File giáo viên": self.teacher_file
        }
        
        for name, path in input_files.items():
            exists = os.path.exists(path)
            size = os.path.getsize(path) if exists else 0
            results["input_files"][name] = {"path": path, "exists": exists, "size": size}
            status = "✅" if exists else "❌"
            size_str = f"({size:,} bytes)" if exists else ""
            print(f"   {name}: {status} {size_str}")
        
        # Kiểm tra file template
        template_files = {
            "Template Excel": self.template_file
        }
        
        for name, path in template_files.items():
            exists = os.path.exists(path)
            size = os.path.getsize(path) if exists else 0
            results["template_files"][name] = {"path": path, "exists": exists, "size": size}
            status = "✅" if exists else "❌"
            size_str = f"({size:,} bytes)" if exists else ""
            print(f"   {name}: {status} {size_str}")
        
        # Kiểm tra file config
        config_files = {
            "Service Account JSON": self.service_account_file,
            "Config Python": self.config_file,
            "Google API Python": self.google_api_file
        }
        
        for name, path in config_files.items():
            exists = os.path.exists(path)
            size = os.path.getsize(path) if exists else 0
            results["config_files"][name] = {"path": path, "exists": exists, "size": size}
            status = "✅" if exists else "❌"
            size_str = f"({size:,} bytes)" if exists else ""
            print(f"   {name}: {status} {size_str}")
        
        return results
    
    def check_google_api_config(self) -> Dict[str, any]:
        """
        Kiểm tra cấu hình Google API
        
        Returns:
            Dict: Kết quả kiểm tra Google API
        """
        print("\n🌐 KIỂM TRA CẤU HÌNH GOOGLE API:")
        print("=" * 40)
        
        results = {
            "service_account_valid": False,
            "google_api_module": False,
            "config_module": False,
            "connection_test": False,
            "errors": []
        }
        
        try:
            # Kiểm tra service account JSON
            if os.path.exists(self.service_account_file):
                try:
                    with open(self.service_account_file, 'r') as f:
                        service_data = json.load(f)
                    
                    required_keys = ['type', 'project_id', 'private_key', 'client_email']
                    missing_keys = [key for key in required_keys if key not in service_data]
                    
                    if not missing_keys:
                        results["service_account_valid"] = True
                        print("   Service Account JSON: ✅ Hợp lệ")
                    else:
                        results["errors"].append(f"Service Account thiếu keys: {missing_keys}")
                        print("   Service Account JSON: ❌ Thiếu thông tin")
                        
                except json.JSONDecodeError:
                    results["errors"].append("Service Account JSON không hợp lệ")
                    print("   Service Account JSON: ❌ Format không hợp lệ")
            else:
                results["errors"].append("Service Account JSON không tồn tại")
                print("   Service Account JSON: ❌ Không tồn tại")
            
            # Kiểm tra module Google API
            try:
                from config.google_api import GoogleAPIClient
                results["google_api_module"] = True
                print("   Google API Module: ✅ Import thành công")
                
                # Test connection
                try:
                    client = GoogleAPIClient()
                    if client.test_connection():
                        results["connection_test"] = True
                        print("   Google API Connection: ✅ Kết nối thành công")
                    else:
                        results["errors"].append("Google API connection test thất bại")
                        print("   Google API Connection: ❌ Kết nối thất bại")
                        
                except Exception as e:
                    results["errors"].append(f"Google API connection error: {str(e)}")
                    print(f"   Google API Connection: ❌ Lỗi: {str(e)}")
                    
            except ImportError as e:
                results["errors"].append(f"Google API module import error: {str(e)}")
                print(f"   Google API Module: ❌ Import thất bại: {str(e)}")
            
            # Kiểm tra config module
            try:
                from config.config import Config
                results["config_module"] = True
                print("   Config Module: ✅ Import thành công")
            except ImportError as e:
                results["errors"].append(f"Config module import error: {str(e)}")
                print(f"   Config Module: ❌ Import thất bại: {str(e)}")
                
        except Exception as e:
            results["errors"].append(f"General error: {str(e)}")
            print(f"   ❌ Lỗi chung: {str(e)}")
        
        return results
    
    def check_data_integrity(self) -> Dict[str, any]:
        """
        Kiểm tra tính toàn vẹn của dữ liệu
        
        Returns:
            Dict: Kết quả kiểm tra dữ liệu
        """
        print("\n📊 KIỂM TRA TÍNH TOÀN VẸN DỮ LIỆU:")
        print("=" * 40)
        
        results = {
            "students": {"count": 0, "valid": False, "errors": []},
            "teachers": {"count": 0, "valid": False, "errors": []},
            "template": {"valid": False, "sheets": [], "errors": []}
        }
        
        # Kiểm tra dữ liệu học sinh
        try:
            if os.path.exists(self.student_file):
                df_students = pd.read_excel(
                    self.student_file, 
                    sheet_name="Danh sách HS toàn trường",
                    header=5,
                    engine='openpyxl'
                )
                df_students = df_students.dropna(how='all')
                
                # Kiểm tra cột STT có phải số không
                valid_students = df_students[pd.to_numeric(df_students.iloc[:, 0], errors='coerce').notna()]
                
                results["students"]["count"] = len(valid_students)
                results["students"]["valid"] = len(valid_students) > 0
                
                if len(valid_students) == 0:
                    results["students"]["errors"].append("Không có học sinh hợp lệ")
                
                print(f"   Dữ liệu học sinh: ✅ {len(valid_students)} học sinh hợp lệ")
                
            else:
                results["students"]["errors"].append("File học sinh không tồn tại")
                print("   Dữ liệu học sinh: ❌ File không tồn tại")
                
        except Exception as e:
            results["students"]["errors"].append(f"Lỗi đọc file học sinh: {str(e)}")
            print(f"   Dữ liệu học sinh: ❌ Lỗi: {str(e)}")
        
        # Kiểm tra dữ liệu giáo viên
        try:
            if os.path.exists(self.teacher_file):
                df_teachers = pd.read_excel(
                    self.teacher_file,
                    sheet_name=0,
                    engine='openpyxl'
                )
                df_teachers = df_teachers.dropna(how='all')
                
                # Kiểm tra cột STT có phải số không
                valid_teachers = df_teachers[pd.to_numeric(df_teachers.iloc[:, 0], errors='coerce').notna()]
                
                results["teachers"]["count"] = len(valid_teachers)
                results["teachers"]["valid"] = len(valid_teachers) > 0
                
                if len(valid_teachers) == 0:
                    results["teachers"]["errors"].append("Không có giáo viên hợp lệ")
                
                print(f"   Dữ liệu giáo viên: ✅ {len(valid_teachers)} giáo viên hợp lệ")
                
            else:
                results["teachers"]["errors"].append("File giáo viên không tồn tại")
                print("   Dữ liệu giáo viên: ❌ File không tồn tại")
                
        except Exception as e:
            results["teachers"]["errors"].append(f"Lỗi đọc file giáo viên: {str(e)}")
            print(f"   Dữ liệu giáo viên: ❌ Lỗi: {str(e)}")
        
        # Kiểm tra template
        try:
            if os.path.exists(self.template_file):
                from openpyxl import load_workbook
                wb = load_workbook(self.template_file)
                
                required_sheets = ['ADMIN', 'GIAO-VIEN', 'HOC-SINH']
                existing_sheets = wb.sheetnames
                
                results["template"]["sheets"] = existing_sheets
                
                missing_sheets = [sheet for sheet in required_sheets if sheet not in existing_sheets]
                
                if not missing_sheets:
                    results["template"]["valid"] = True
                    print(f"   Template Excel: ✅ Có đủ {len(required_sheets)} sheet bắt buộc")
                else:
                    results["template"]["errors"].append(f"Thiếu sheets: {missing_sheets}")
                    print(f"   Template Excel: ❌ Thiếu sheets: {missing_sheets}")
                
            else:
                results["template"]["errors"].append("File template không tồn tại")
                print("   Template Excel: ❌ File không tồn tại")
                
        except Exception as e:
            results["template"]["errors"].append(f"Lỗi đọc template: {str(e)}")
            print(f"   Template Excel: ❌ Lỗi: {str(e)}")
        
        return results
    
    def check_python_dependencies(self) -> Dict[str, bool]:
        """
        Kiểm tra các Python dependencies
        
        Returns:
            Dict: Kết quả kiểm tra dependencies
        """
        print("\n🐍 KIỂM TRA PYTHON DEPENDENCIES:")
        print("=" * 40)
        
        required_packages = {
            'pandas': 'pandas',
            'openpyxl': 'openpyxl', 
            'google-auth': 'google.auth',
            'google-auth-oauthlib': 'google_auth_oauthlib',
            'google-auth-httplib2': 'google_auth_httplib2',
            'google-api-python-client': 'googleapiclient',
            'gspread': 'gspread'
        }
        
        results = {}
        
        for package_name, import_name in required_packages.items():
            try:
                __import__(import_name)
                results[package_name] = True
                print(f"   {package_name}: ✅ Đã cài đặt")
            except ImportError:
                results[package_name] = False
                print(f"   {package_name}: ❌ Chưa cài đặt")
        
        return results
    
    def run_full_check(self) -> Dict[str, any]:
        """
        Chạy kiểm tra toàn diện hệ thống
        
        Returns:
            Dict: Kết quả kiểm tra tổng hợp
        """
        print("\n🔍 KIỂM TRA TOÀN DIỆN HỆ THỐNG")
        print("=" * 60)
        
        results = {
            "file_system": self.check_file_system(),
            "google_api": self.check_google_api_config(),
            "data_integrity": self.check_data_integrity(),
            "python_dependencies": self.check_python_dependencies(),
            "overall_status": "unknown"
        }
        
        # Đánh giá trạng thái tổng thể
        critical_issues = []
        
        # Kiểm tra file system
        if not all(results["file_system"]["directories"][d]["exists"] for d in results["file_system"]["directories"]):
            critical_issues.append("Thiếu thư mục quan trọng")
        
        if not all(results["file_system"]["input_files"][f]["exists"] for f in results["file_system"]["input_files"]):
            critical_issues.append("Thiếu file input")
        
        if not results["file_system"]["template_files"]["Template Excel"]["exists"]:
            critical_issues.append("Thiếu file template")
        
        # Kiểm tra dữ liệu
        if not results["data_integrity"]["students"]["valid"]:
            critical_issues.append("Dữ liệu học sinh không hợp lệ")
        
        if not results["data_integrity"]["teachers"]["valid"]:
            critical_issues.append("Dữ liệu giáo viên không hợp lệ")
        
        if not results["data_integrity"]["template"]["valid"]:
            critical_issues.append("Template không hợp lệ")
        
        # Kiểm tra dependencies
        core_packages = ['pandas', 'openpyxl']
        if not all(results["python_dependencies"][pkg] for pkg in core_packages):
            critical_issues.append("Thiếu Python packages quan trọng")
        
        # Đánh giá tổng thể
        if not critical_issues:
            if results["google_api"]["connection_test"]:
                results["overall_status"] = "excellent"
            elif results["google_api"]["service_account_valid"]:
                results["overall_status"] = "good"
            else:
                results["overall_status"] = "fair"
        else:
            results["overall_status"] = "poor"
        
        results["critical_issues"] = critical_issues
        
        # In kết quả tổng kết
        print(f"\n📋 KẾT QUẢ TỔNG KẾT:")
        print("=" * 40)
        
        status_icons = {
            "excellent": "🟢",
            "good": "🟡", 
            "fair": "🟠",
            "poor": "🔴"
        }
        
        status_texts = {
            "excellent": "XUẤT SẮC - Tất cả chức năng hoạt động tốt",
            "good": "TỐT - Chức năng cơ bản hoạt động, Google API cần kiểm tra",
            "fair": "TRUNG BÌNH - Chức năng cơ bản hoạt động, thiếu Google API",
            "poor": "YẾU - Có vấn đề nghiêm trọng cần khắc phục"
        }
        
        status = results["overall_status"]
        print(f"   Trạng thái tổng thể: {status_icons[status]} {status_texts[status]}")
        
        if critical_issues:
            print(f"\n   ⚠️  Vấn đề cần khắc phục:")
            for issue in critical_issues:
                print(f"      - {issue}")
        
        return results
