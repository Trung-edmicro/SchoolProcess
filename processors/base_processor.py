"""
Base Data Processor
Class cơ sở cho tất cả các processor
Author: Assistant
Date: 2025-07-26
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
import os
import re
from datetime import datetime
from typing import Dict, List, Any, Optional
from abc import ABC, abstractmethod


class BaseDataProcessor(ABC):
    """Base class cho tất cả các data processor"""
    
    def __init__(self, input_folder: str, temp_folder: str, output_folder: str = None):
        """
        Khởi tạo BaseDataProcessor
        
        Args:
            input_folder (str): Thư mục chứa file input
            temp_folder (str): Thư mục chứa template
            output_folder (str): Thư mục output (mặc định là temp_folder)
        """
        self.input_folder = input_folder
        self.temp_folder = temp_folder
        self.output_folder = output_folder if output_folder else temp_folder
        
        # Đường dẫn các file
        self.student_file = os.path.join(input_folder, "Danh sach hoc sinh.xlsx")
        self.teacher_file = os.path.join(input_folder, "DS tài khoản giáo viên.xlsx")
        self.template_file = os.path.join(temp_folder, "Template_Export.xlsx")
        
        # Dữ liệu
        self.school_name = ""
        self.students_data = None
        self.teachers_data = None
        self.template_wb = None
    
    def extract_school_name(self) -> str:
        """
        Trích xuất tên trường từ file học sinh
        
        Returns:
            str: Tên trường
        """
        try:
            # Đọc hàng đầu tiên từ file học sinh
            df = pd.read_excel(self.student_file, sheet_name="Danh sách HS toàn trường", header=None, nrows=1)
            
            # Lấy giá trị đầu tiên (tên trường)
            school_title = str(df.iloc[0, 0])
            
            # Trích xuất tên trường từ chuỗi
            # Format: "Danh sách học sinh trường: TRUNG TÂM GDNN - GDTX TP CHÍ LINH - năm học: 2024 - 2025"
            match = re.search(r'trường:\s*(.+?)\s*-\s*năm học', school_title, re.IGNORECASE)
            if match:
                school_name = match.group(1).strip()
            else:
                # Fallback: lấy phần sau "trường:"
                parts = school_title.split("trường:")
                if len(parts) > 1:
                    school_name = parts[1].split("-")[0].strip()
                else:
                    school_name = "TRƯỜNG KHÔNG XÁC ĐỊNH"
            
            self.school_name = school_name
            print(f"✅ Đã trích xuất tên trường: {school_name}")
            return school_name
            
        except Exception as e:
            print(f"❌ Lỗi khi trích xuất tên trường: {e}")
            self.school_name = "TRƯỜNG KHÔNG XÁC ĐỊNH"
            return self.school_name
    
    def load_template(self) -> openpyxl.Workbook:
        """
        Load template workbook
        
        Returns:
            openpyxl.Workbook: Template workbook
        """
        try:
            self.template_wb = load_workbook(self.template_file)
            print(f"✅ Đã load template: {self.template_file}")
            return self.template_wb
            
        except Exception as e:
            print(f"❌ Lỗi khi load template: {e}")
            return None
    
    def apply_border_to_sheet(self, sheet, max_row: int, max_col: int, center_columns: List[int] = None):
        """
        Áp dụng border và căn giữa cho các ô có dữ liệu trong sheet
        
        Args:
            sheet: Sheet cần áp dụng border
            max_row (int): Số hàng tối đa có dữ liệu
            max_col (int): Số cột tối đa có dữ liệu
            center_columns (List[int]): Danh sách các cột cần căn giữa (1-based)
        """
        try:
            # Định nghĩa style border đen
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Định nghĩa alignment căn giữa
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            if center_columns is None:
                center_columns = []
            
            # Áp dụng border cho tất cả các ô có dữ liệu
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    # Chỉ áp dụng border cho ô có dữ liệu hoặc là header
                    if cell.value is not None or row == 1:
                        cell.border = thin_border
                        
                        # Căn giữa cho các cột được chỉ định
                        if col in center_columns:
                            cell.alignment = center_alignment
            
            print(f"✅ Đã áp dụng border và căn giữa cho sheet {sheet.title}")
            
        except Exception as e:
            print(f"❌ Lỗi khi áp dụng style cho sheet {sheet.title}: {e}")
    
    def update_admin_sheet(self):
        """Cập nhật sheet ADMIN với tên trường"""
        try:
            if not self.template_wb:
                self.load_template()
            
            admin_sheet = self.template_wb['ADMIN']
            
            # Cập nhật tên trường tại ô A1
            admin_sheet['A1'] = f"{self.school_name}"
            
            print(f"✅ Đã cập nhật sheet ADMIN với tên trường: {self.school_name}")
            
        except Exception as e:
            print(f"❌ Lỗi khi cập nhật sheet ADMIN: {e}")
    
    def update_teachers_sheet(self):
        """Cập nhật sheet GIAO-VIEN với dữ liệu giáo viên"""
        try:
            if not self.template_wb:
                self.load_template()
            
            if self.teachers_data is None:
                self.load_teachers_data()
            
            teacher_sheet = self.template_wb['GIAO-VIEN']
            
            # Xóa dữ liệu cũ (giữ lại header)
            for row in range(2, teacher_sheet.max_row + 1):
                for col in range(1, teacher_sheet.max_column + 1):
                    teacher_sheet.cell(row=row, column=col).value = None
            
            # Ghi dữ liệu mới
            for idx, (_, row_data) in enumerate(self.teachers_data.iterrows(), start=2):
                # STT
                teacher_sheet.cell(row=idx, column=1).value = row_data.get('STT', idx-1)
                # Tên giáo viên
                teacher_sheet.cell(row=idx, column=2).value = row_data.get('Tên giáo viên', '')
                # Ngày sinh
                teacher_sheet.cell(row=idx, column=3).value = row_data.get('Ngày sinh', '')
                # Tên đăng nhập
                teacher_sheet.cell(row=idx, column=4).value = row_data.get('Tên đăng nhập', '')
                # Mật khẩu đăng nhập lần đầu
                teacher_sheet.cell(row=idx, column=5).value = row_data.get('Mật khẩu đăng nhập lần đầu', '')
            
            # Áp dụng border và căn giữa cho các ô có dữ liệu
            max_data_row = len(self.teachers_data) + 1  # +1 cho header
            max_data_col = 5  # 5 cột: STT, Tên, Ngày sinh, Tên đăng nhập, Mật khẩu
            
            # Các cột cần căn giữa: STT (1), Ngày sinh (3), Mật khẩu (5)
            center_columns = [1, 3, 5]
            self.apply_border_to_sheet(teacher_sheet, max_data_row, max_data_col, center_columns)
            
            print(f"✅ Đã cập nhật sheet GIAO-VIEN với {len(self.teachers_data)} giáo viên")
            
        except Exception as e:
            print(f"❌ Lỗi khi cập nhật sheet GIAO-VIEN: {e}")
    
    def update_students_sheet(self):
        """Cập nhật sheet HOC-SINH với dữ liệu học sinh"""
        try:
            if not self.template_wb:
                self.load_template()
            
            if self.students_data is None:
                self.load_students_data()
            
            student_sheet = self.template_wb['HOC-SINH']
            
            # Xóa dữ liệu cũ (giữ lại header)
            for row in range(2, student_sheet.max_row + 1):
                for col in range(1, student_sheet.max_column + 1):
                    student_sheet.cell(row=row, column=col).value = None
            
            # Ghi dữ liệu mới
            for idx, (_, row_data) in enumerate(self.students_data.iterrows(), start=2):
                # STT
                student_sheet.cell(row=idx, column=1).value = row_data.get('STT', idx-1)
                # Họ và tên
                student_sheet.cell(row=idx, column=2).value = row_data.get('Họ và tên', '')
                # Ngày sinh
                student_sheet.cell(row=idx, column=3).value = row_data.get('Ngày sinh', '')
                # Lớp chính
                student_sheet.cell(row=idx, column=4).value = row_data.get('Lớp chính', '')
                # Tài khoản
                student_sheet.cell(row=idx, column=5).value = row_data.get('Tài khoản', '')
                # Mật khẩu lần đầu
                student_sheet.cell(row=idx, column=6).value = row_data.get('Mật khẩu lần đầu', '')
                # Mã đăng nhập cho PH
                student_sheet.cell(row=idx, column=7).value = row_data.get('Mã đăng nhập cho PH', '')
            
            # Áp dụng border và căn giữa cho các ô có dữ liệu
            max_data_row = len(self.students_data) + 1  # +1 cho header
            max_data_col = 7  # 7 cột: STT, Họ tên, Ngày sinh, Lớp, Tài khoản, Mật khẩu, Mã PH
            
            # Các cột cần căn giữa: STT (1), Ngày sinh (3), Lớp (4), Mật khẩu (6), Mã PH (7)
            center_columns = [1, 3, 4, 6, 7]
            self.apply_border_to_sheet(student_sheet, max_data_row, max_data_col, center_columns)
            
            print(f"✅ Đã cập nhật sheet HOC-SINH với {len(self.students_data)} học sinh")
            
        except Exception as e:
            print(f"❌ Lỗi khi cập nhật sheet HOC-SINH: {e}")
    
    def save_output_file(self) -> str:
        """
        Lưu file output với tên theo format
        
        Returns:
            str: Đường dẫn file output
        """
        try:
            if not self.school_name:
                self.extract_school_name()
            
            # Tạo tên file output
            safe_school_name = re.sub(r'[^\w\s-]', '', self.school_name)
            safe_school_name = re.sub(r'[-\s]+', '_', safe_school_name)
            output_filename = f"Export_{safe_school_name}.xlsx"
            output_path = os.path.join(self.output_folder, output_filename)
            
            # Lưu file local
            if self.template_wb:
                self.template_wb.save(output_path)
                print(f"✅ Đã lưu file output: {output_path}")
                
                # Thực hiện post-save actions (có thể override trong subclass)
                self._post_save_actions(output_path, output_filename)
                
                return output_path
            else:
                raise ValueError("Template workbook chưa được load")
                
        except Exception as e:
            print(f"❌ Lỗi khi lưu file output: {e}")
            return ""
    
    def _post_save_actions(self, output_path: str, output_filename: str):
        """
        Thực hiện các hành động sau khi lưu file (có thể override trong subclass)
        
        Args:
            output_path (str): Đường dẫn file đã lưu
            output_filename (str): Tên file
        """
        pass  # Base implementation không làm gì
    
    def print_summary(self):
        """In tóm tắt quá trình xử lý"""
        print(f"\n📊 TÓM TẮT QUÁ TRÌNH XỬ LÝ:")
        print(f"   🏫 Tên trường: {self.school_name}")
        print(f"   👨‍🎓 Số học sinh: {len(self.students_data) if self.students_data is not None else 0}")
        print(f"   👨‍🏫 Số giáo viên: {len(self.teachers_data) if self.teachers_data is not None else 0}")
        print(f"   📁 Thư mục input: {self.input_folder}")
        print(f"   📁 Thư mục output: {self.output_folder}")
    
    def process_all(self) -> str:
        """
        Xử lý toàn bộ quy trình mapping
        
        Returns:
            str: Đường dẫn file output
        """
        print(f"\n🚀 BẮT ĐẦU QUY TRÌNH MAPPING DỮ LIỆU - {self.get_processor_name()}")
        print("=" * 60)
        
        try:
            # 1. Trích xuất tên trường
            print("\n📋 Bước 1: Trích xuất tên trường...")
            self.extract_school_name()
            
            # 2. Load dữ liệu
            print("\n📋 Bước 2: Load dữ liệu input...")
            self.load_students_data()
            self.load_teachers_data()
            
            # 3. Load template
            print("\n📋 Bước 3: Load template...")
            self.load_template()
            
            # 4. Cập nhật các sheet
            print("\n📋 Bước 4: Mapping dữ liệu...")
            self.update_admin_sheet()
            self.update_teachers_sheet()
            self.update_students_sheet()
            
            # 5. Lưu file output
            print("\n📋 Bước 5: Lưu file output...")
            output_path = self.save_output_file()
            
            print(f"\n✅ HOÀN THÀNH! File output: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"\n❌ LỖI TRONG QUY TRÌNH: {e}")
            return ""
    
    # Abstract methods - phải implement trong subclass
    @abstractmethod
    def load_students_data(self) -> pd.DataFrame:
        """Load dữ liệu học sinh - phải implement trong subclass"""
        pass
    
    @abstractmethod  
    def load_teachers_data(self) -> pd.DataFrame:
        """Load dữ liệu giáo viên - phải implement trong subclass"""
        pass
    
    @abstractmethod
    def get_processor_name(self) -> str:
        """Trả về tên processor - phải implement trong subclass"""
        pass
