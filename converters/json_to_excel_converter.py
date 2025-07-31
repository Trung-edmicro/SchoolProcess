"""
JSON to Excel Template Converter (Chuẩn Mode 1)
Chuyển đổi dữ liệu JSON từ OnLuyen API thành file Excel format đúng template Mode 1
Author: Assistant  
Date: 2025-07-26
"""

import json
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from datetime import datetime
import os
import shutil
from pathlib import Path


class JSONToExcelTemplateConverter:
    """Converter chuyển JSON sang Excel format theo template chuẩn Mode 1"""
    
    def __init__(self, json_file_path: str, template_path: str = None):
        """
        Khởi tạo converter
        
        Args:
            json_file_path (str): Đường dẫn file JSON
            template_path (str): Đường dẫn template Excel
        """
        self.json_file_path = json_file_path
        self.template_path = template_path or "data/temp/Template_Export.xlsx"
        self.json_data = None
        self.school_name = ""
        self.admin_email = ""
        self.admin_password = ""  # Thêm biến lưu admin password
        self.teachers_df = None
        self.students_df = None
        
        # Thiết lập style cho Excel (chuẩn như base_processor)
        self.thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'), 
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Alignment styles
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Font styles
        self.header_font = Font(bold=True, size=11, name='Calibri')
        self.data_font = Font(size=11, name='Calibri')
        
        # Fill styles
        self.header_fill = PatternFill(start_color="D9EDF7", end_color="D9EDF7", fill_type="solid")
        
    def load_json_data(self):
        """Load dữ liệu từ file JSON"""
        try:
            with open(self.json_file_path, 'r', encoding='utf-8') as f:
                self.json_data = json.load(f)
            
            print(f"✅ Đã load JSON data từ: {self.json_file_path}")
            
            # Extract school info từ cấu trúc mới (có thể có metadata hoặc school_info trực tiếp)
            if 'metadata' in self.json_data:
                # Cấu trúc cũ với metadata
                metadata = self.json_data.get('metadata', {})
                school_info = metadata.get('school_info', {})
                self.admin_password = metadata.get('admin_password', '123456')
            else:
                # Cấu trúc mới với school_info trực tiếp
                school_info = self.json_data.get('school_info', {})
                # Sửa: password được lưu trong school_info.password, không phải admin_password
                self.admin_password = school_info.get('password', '123456')
            
            self.school_name = school_info.get('name', 'Unknown School')
            self.admin_email = school_info.get('admin', '')
            
            print(f"📋 Tên trường: {self.school_name}")
            print(f"📧 Admin email: {self.admin_email}")
            print(f"🔑 Admin password: {'*' * len(self.admin_password) if self.admin_password else 'N/A'}")
            
            return True
            
        except Exception as e:
            print(f"❌ Lỗi khi load JSON: {e}")
            return False
    
    def extract_teachers_data(self):
        """Trích xuất dữ liệu giáo viên từ JSON - hỗ trợ cả workflow và filtered format"""
        try:
            print("   📊 Đang trích xuất dữ liệu giáo viên...")
            
            # Xử lý các format khác nhau của dữ liệu teachers
            teachers_raw = self.json_data.get('teachers', [])
            
            if isinstance(teachers_raw, dict):
                # Format workflow thường: {'data': [...]}
                teachers_data = teachers_raw.get('data', [])
                print(f"   📋 Phát hiện format workflow: {len(teachers_data)} giáo viên")
            elif isinstance(teachers_raw, list):
                # Format filtered: [...] trực tiếp
                teachers_data = teachers_raw
                print(f"   📋 Phát hiện format filtered: {len(teachers_data)} giáo viên")
            else:
                print(f"   ❌ Format dữ liệu teachers không được hỗ trợ: {type(teachers_raw)}")
                return False
            
            if not teachers_data:
                print("   ⚠️ Không có dữ liệu giáo viên")
                return False
            
            teachers_list = []
            
            stt = 1  # Đếm STT riêng để không bị lỗ khi loại bỏ GVCN
            for teacher_record in teachers_data:
                # Xử lý cấu trúc teacher record linh hoạt
                if isinstance(teacher_record, dict):
                    # Tìm teacherInfo trong các vị trí có thể
                    teacher_info_data = (
                        teacher_record.get('teacherInfo', {}) or 
                        teacher_record.get('userInfo', {}) or 
                        teacher_record
                    )
                    
                    teacher_name = teacher_info_data.get('displayName', '').strip()
                    
                    # Bỏ qua giáo viên có tên là "GVCN"
                    if teacher_name.upper() == "GVCN":
                        print(f"   🚫 Loại bỏ giáo viên: {teacher_name} (GVCN)")
                        continue
                    
                    teacher_info = {
                        'STT': stt,
                        'Tên giáo viên': teacher_name,
                        'Ngày sinh': teacher_info_data.get('userBirthday', ''),
                        'Tên đăng nhập': teacher_info_data.get('userName', ''),
                        'Mật khẩu đăng nhập lần đầu': teacher_info_data.get('pwd', '')
                    }
                    teachers_list.append(teacher_info)
                    stt += 1
            
            self.teachers_df = pd.DataFrame(teachers_list)
            print(f"✅ Đã trích xuất {len(teachers_list)} giáo viên")
            
            return True
            
        except Exception as e:
            print(f"❌ Lỗi khi trích xuất dữ liệu giáo viên: {e}")
            return False
    
    def extract_students_data(self):
        """Trích xuất dữ liệu học sinh từ JSON - hỗ trợ cả workflow và filtered format"""
        try:
            print("   📊 Đang trích xuất dữ liệu học sinh...")
            
            # Xử lý các format khác nhau của dữ liệu students
            students_raw = self.json_data.get('students', [])
            
            if isinstance(students_raw, dict):
                # Format workflow thường: {'data': [...]}
                students_data = students_raw.get('data', [])
                print(f"   📋 Phát hiện format workflow: {len(students_data)} học sinh")
            elif isinstance(students_raw, list):
                # Format filtered: [...] trực tiếp
                students_data = students_raw
                print(f"   📋 Phát hiện format filtered: {len(students_data)} học sinh")
            else:
                print(f"   ❌ Format dữ liệu students không được hỗ trợ: {type(students_raw)}")
                return False
            
            if not students_data:
                print("   ⚠️ Không có dữ liệu học sinh")
                return False
            
            students_list = []
            
            for idx, student_record in enumerate(students_data, 1):
                # Xử lý cấu trúc student record linh hoạt
                if isinstance(student_record, dict):
                    # Tìm userInfo trong các vị trí có thể
                    user_info = (
                        student_record.get('userInfo', {}) or 
                        student_record.get('studentInfo', {}) or 
                        student_record
                    )
                    
                    groupClass = student_record.get('groupClass', [])

                    student_info = {
                        'STT': idx,
                        'Họ và tên': user_info.get('displayName', ''),
                        'Ngày sinh': user_info.get('userBirthday', ''),
                        'Khối': student_record.get('grade', ''),
                        'Lớp': groupClass[0].get('className', '') if groupClass else '',
                        'Tài khoản': user_info.get('userName', ''),
                        'Mật khẩu lần đầu': user_info.get('pwd', ''),
                        'Mã đăng nhập cho PH': user_info.get('codePin', '')
                    }
                    students_list.append(student_info)
            
            self.students_df = pd.DataFrame(students_list)
            print(f"✅ Đã trích xuất {len(students_list)} học sinh")
            
            return True
            
        except Exception as e:
            print(f"❌ Lỗi khi trích xuất dữ liệu học sinh: {e}")
            return False
    
    def apply_border_to_sheet(self, sheet, max_row: int, max_col: int, center_columns: list = None):
        """
        Áp dụng border và căn giữa cho các ô có dữ liệu trong sheet (chuẩn như base_processor)
        
        Args:
            sheet: Sheet cần áp dụng border
            max_row (int): Số hàng tối đa có dữ liệu
            max_col (int): Số cột tối đa có dữ liệu
            center_columns (list): Danh sách các cột cần căn giữa (1-based)
        """
        try:
            if center_columns is None:
                center_columns = []
            
            # Áp dụng border cho tất cả các ô có dữ liệu
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    # Chỉ áp dụng border cho ô có dữ liệu hoặc là header
                    if cell.value is not None or row == 1:
                        cell.border = self.thin_border
                        
                        # Căn giữa cho các cột được chỉ định
                        if col in center_columns:
                            cell.alignment = self.center_alignment
                        else:
                            cell.alignment = self.left_alignment
            
            print(f"✅ Đã áp dụng border và căn giữa cho sheet {sheet.title}")
            
        except Exception as e:
            print(f"❌ Lỗi khi áp dụng style cho sheet {sheet.title}: {e}")

    def copy_template(self, output_path: str):
        """Copy template làm base cho file output"""
        try:
            shutil.copy2(self.template_path, output_path)
            print(f"✅ Đã copy template thành: {output_path}")
            return True
        except Exception as e:
            print(f"❌ Lỗi khi copy template: {e}")
            return False
    
    def update_admin_sheet(self, workbook):
        """Cập nhật sheet ADMIN với thông tin trường, admin và HT/HP - PRESERVE MERGED CELLS"""
        try:
            admin_sheet = workbook['ADMIN']
            
            print(f"🔍 DEBUG: Template structure analysis...")
            # Phân tích merged cells hiện có
            merged_ranges = []
            if admin_sheet.merged_cells:
                merged_ranges = list(admin_sheet.merged_cells.ranges)
                print(f"   📋 Merged cells found: {[str(r) for r in merged_ranges]}")
            
            # GIỮ NGUYÊN merged cells - KHÔNG unmerge
            # Chỉ cập nhật nội dung các ô merged chính
            
            # 1. Cập nhật A1 (merged A1:D1) - Tên trường
            admin_sheet['A1'] = f"{self.school_name}"
            admin_sheet['A1'].font = Font(bold=True, size=14, name='Calibri')
            admin_sheet['A1'].alignment = self.center_alignment
            print(f"   ✅ A1 (merged A1:D1): Tên trường")
            
            # 2. Headers đã có sẵn trong template: C2="Tài khoản", D2="Mật khẩu lần đầu"
            # Chỉ format lại headers nếu cần
            try:
                if admin_sheet['C2'].value:
                    admin_sheet['C2'].font = self.header_font
                    admin_sheet['C2'].alignment = self.center_alignment
                if admin_sheet['D2'].value:
                    admin_sheet['D2'].font = self.header_font
                    admin_sheet['D2'].alignment = self.center_alignment
                print(f"   ✅ Headers C2, D2: Đã format")
            except:
                pass
            
            # 3. Admin data (Row 3: A3:B3 merged = "Admin")
            # CHỈ điền data vào C3, D3 - KHÔNG touch B3 vì đó là merged cell secondary
            admin_sheet['C3'] = self.admin_email
            admin_sheet['D3'] = self.admin_password
            
            # Format admin row (chỉ C3, D3)
            for col in ['C', 'D']:
                cell = admin_sheet[f'{col}3']
                cell.font = self.data_font
                if col == 'C':  # Tài khoản left align
                    cell.alignment = self.left_alignment
                else:  # Mật khẩu center align
                    cell.alignment = self.center_alignment
            
            print(f"   ✅ Row 3 (Admin): A3:B3=merged, C3={self.admin_email}, D3=***")
            
            # 4. HT/HP data từ JSON
            ht_hp_info = self.json_data.get('ht_hp_info', {})
            ht_list = ht_hp_info.get('ht', [])
            hp_list = ht_hp_info.get('hp', [])
            
            accounts_filled = 1  # Admin đã điền
            
            # 5. Hiệu Trưởng (Row 4)
            if ht_list:
                ht = ht_list[0]  # Chỉ lấy HT đầu tiên để fit template
                # Kiểm tra nếu B4 không nằm trong merged cell thì mới điền tên
                try:
                    admin_sheet['B4'] = ht.get('displayName', '')  # Tên HT vào cột B4
                    print(f"   ✅ B4: Điền tên HT = {ht.get('displayName', '')}")
                except:
                    print(f"   ⚠️ B4: Bị merged, skip điền tên HT")
                
                admin_sheet['C4'] = ht.get('userName', '')
                admin_sheet['D4'] = ht.get('pwd', '')
                
                # Format HT row (chỉ format các ô không merged)
                for col in ['C', 'D']:  # Chắc chắn C4, D4 không merged
                    cell = admin_sheet[f'{col}4']
                    cell.font = self.data_font
                    if col == 'C':  # Username left align
                        cell.alignment = self.left_alignment
                    else:  # Mật khẩu center align
                        cell.alignment = self.center_alignment
                
                # Format B4 nếu điền được
                try:
                    admin_sheet['B4'].font = self.data_font
                    admin_sheet['B4'].alignment = self.center_alignment
                except:
                    pass
                
                accounts_filled += 1
                print(f"   ✅ Row 4 (Hiệu Trưởng): C4={ht.get('userName', '')}, D4=***")
                
                if len(ht_list) > 1:
                    print(f"   ⚠️ Template chỉ hỗ trợ 1 HT, có {len(ht_list)} HT")
            else:
                print(f"   📋 Row 4 (Hiệu Trưởng): Không có dữ liệu")
            
            # 6. Hiệu Phó (Row 5 và các row tiếp theo nếu có nhiều HP)
            if hp_list:
                # BACKUP row 7 content trước khi thực hiện insert (để preserve "Lưu ý")
                row7_backup = {
                    'A7': admin_sheet['A7'].value,
                    'merged_cell_range': None
                }
                
                # Tìm merged cell chứa row 7
                for merged_range in admin_sheet.merged_cells.ranges:
                    if 'A7' in str(merged_range) or any(f'{chr(65+i)}7' in str(merged_range) for i in range(4)):
                        row7_backup['merged_cell_range'] = str(merged_range)
                        break
                
                print(f"   🔒 Backup row 7: A7='{row7_backup['A7']}', merged='{row7_backup['merged_cell_range']}'")
                
                # Xử lý HP đầu tiên vào row 5 (có sẵn trong template)
                first_hp = hp_list[0]
                admin_sheet['B5'] = first_hp.get('displayName', '')
                admin_sheet['C5'] = first_hp.get('userName', '')
                admin_sheet['D5'] = first_hp.get('pwd', '')
                
                # Format HP đầu tiên
                for col in ['C', 'D']:
                    cell = admin_sheet[f'{col}5']
                    cell.font = self.data_font
                    if col == 'C':
                        cell.alignment = self.left_alignment
                    else:
                        cell.alignment = self.center_alignment
                
                try:
                    admin_sheet['B5'].font = self.data_font
                    admin_sheet['B5'].alignment = self.center_alignment
                except:
                    pass
                
                accounts_filled += 1
                print(f"   ✅ Row 5 (Hiệu Phó 1): B5={first_hp.get('displayName', '')}, C5={first_hp.get('userName', '')}, D5=***")
                
                # Xử lý các HP còn lại - INSERT từ row 6 trở đi
                for idx in range(1, len(hp_list)):
                    hp = hp_list[idx]
                    insert_position = 5 + idx  # Row 6, 7, 8...
                    
                    # Insert row mới
                    admin_sheet.insert_rows(insert_position)
                    print(f"   ➕ Insert row {insert_position} cho HP thứ {idx + 1}")
                    
                    # Điền data cho row mới
                    admin_sheet[f'A{insert_position}'] = "Hiệu Phó"
                    admin_sheet[f'B{insert_position}'] = hp.get('displayName', '')
                    admin_sheet[f'C{insert_position}'] = hp.get('userName', '')
                    admin_sheet[f'D{insert_position}'] = hp.get('pwd', '')
                    
                    # Format row mới
                    admin_sheet[f'A{insert_position}'].font = self.data_font
                    admin_sheet[f'A{insert_position}'].alignment = self.center_alignment  # Căn giữa cho "Hiệu Phó"
                    admin_sheet[f'A{insert_position}'].border = self.thin_border  # Thêm border
                    
                    for col in ['B', 'C', 'D']:
                        cell = admin_sheet[f'{col}{insert_position}']
                        cell.font = self.data_font
                        cell.border = self.thin_border  # Thêm border cho tất cả ô
                        if col == 'B':
                            cell.alignment = self.center_alignment  # Tên căn giữa
                        elif col == 'C':
                            cell.alignment = self.left_alignment    # Username căn trái
                        else:  # col == 'D'
                            cell.alignment = self.center_alignment  # Password căn giữa
                    
                    # Set row height đồng bộ với các row khác (20)
                    admin_sheet.row_dimensions[insert_position].height = 20
                    print(f"   📏 Set row {insert_position} height = 20")
                    
                    accounts_filled += 1
                    print(f"   ✅ Row {insert_position} (Hiệu Phó {idx + 1}): B{insert_position}={hp.get('displayName', '')}, C{insert_position}={hp.get('userName', '')}, D{insert_position}=***")
                
                # QUAN TRỌNG: Reset row height của dòng 7 cũ (bây giờ trống) về bình thường
                try:
                    admin_sheet.row_dimensions[7].height = 20
                    print(f"   📏 Reset row 7 height = 20 (dòng cũ đã trống)")
                except:
                    pass
                
                # RESTORE row 7 content nếu bị mất
                current_row7_value = admin_sheet['A7'].value
                if not current_row7_value and row7_backup['A7']:
                    print(f"   🔧 Row 7 bị mất content, đang restore...")
                    # Tìm vị trí mới của row 7 (có thể đã shift)
                    target_row = 7 + len(hp_list) - 1  # Row 7 gốc + số HP insert
                    
                    # Restore content vào target row
                    admin_sheet[f'A{target_row}'] = row7_backup['A7']
                    
                    # QUAN TRỌNG: Merge cells A{target_row}:D{target_row} để giống template
                    try:
                        # Kiểm tra nếu range chưa được merge
                        target_range = f'A{target_row}:D{target_row}'
                        
                        # Unmerge range cũ nếu tồn tại (có thể bị lệch)
                        ranges_to_remove = []
                        for merged_range in admin_sheet.merged_cells.ranges:
                            range_str = str(merged_range)
                            # Nếu có merged range từ A7 trở đi, remove để tạo mới
                            if range_str.startswith('A7:') or range_str.startswith('A8:') or range_str.startswith('A9:'):
                                ranges_to_remove.append(merged_range)
                        
                        for range_to_remove in ranges_to_remove:
                            admin_sheet.unmerge_cells(str(range_to_remove))
                            print(f"   🔧 Unmerged old range: {range_to_remove}")
                        
                        # Merge cells mới cho "Lưu ý"
                        admin_sheet.merge_cells(target_range)
                        print(f"   ✅ Merged cells: {target_range}")
                        
                        # Format merged cell với wrap text và row height hợp lý
                        target_cell = admin_sheet[f'A{target_row}']
                        target_cell.font = self.data_font
                        target_cell.alignment = Alignment(
                            horizontal='left', 
                            vertical='center',  # Middle align (center) thay vì top
                            wrap_text=True  # Wrap text để xuống dòng
                        )
                        
                        admin_sheet.row_dimensions[target_row].height = 60
                        
                    except Exception as merge_error:
                        print(f"   ⚠️ Lỗi merge cells: {merge_error}")
                    
                    print(f"   ✅ Restored row {target_row} with merged cells: '{row7_backup['A7'][:50]}...'")
                else:
                    print(f"   ✅ Row 7 content vẫn còn nguyên: '{current_row7_value[:50] if current_row7_value else 'Empty'}...'")
                
                if len(hp_list) > 1:
                    print(f"   ✅ Đã insert {len(hp_list) - 1} dòng mới cho tổng {len(hp_list)} Hiệu phó")
            else:
                print(f"   📋 Row 5 (Hiệu Phó): Không có dữ liệu")
            
            # 7. Điều chỉnh column widths một cách an toàn
            try:
                admin_sheet.column_dimensions['A'].width = 15  # Vai trò (merged)
                admin_sheet.column_dimensions['B'].width = 25  # Tên người
                admin_sheet.column_dimensions['C'].width = 35  # Tài khoản
                admin_sheet.column_dimensions['D'].width = 25  # Mật khẩu
            except:
                pass
            
            # 8. Row heights - cập nhật cho tất cả rows có thể có (bao gồm các HP được thêm)
            try:
                admin_sheet.row_dimensions[1].height = 25  # Title
                # Set height cho admin, HT, HP rows (2-6 và các row được insert)
                total_account_rows = 2 + len(ht_list) + len(hp_list)  # Admin + HT + HP
                for row in range(2, total_account_rows + 2):  # +2 để bao gồm đủ rows
                    if row <= 10:  # Giới hạn an toàn
                        admin_sheet.row_dimensions[row].height = 20
                print(f"   📏 Set row heights (2-{min(total_account_rows + 1, 10)}) = 20")
            except:
                pass
            
            print(f"✅ Đã cập nhật sheet ADMIN với {accounts_filled} tài khoản")
            print(f"   📋 1 Admin + {len(ht_list)} HT + {len(hp_list)} HP")
            print(f"   🏗️ PRESERVE merged cells: {len(merged_ranges)} ranges")
            return True
            
        except Exception as e:
            print(f"❌ Lỗi khi cập nhật sheet ADMIN: {e}")
            return False
    
    def fill_teachers_sheet(self, workbook):
        """Điền dữ liệu giáo viên vào sheet GIAO-VIEN"""
        try:
            if self.teachers_df is None or self.teachers_df.empty:
                print("⚠️ Không có dữ liệu giáo viên để điền")
                return True
                
            teachers_sheet = workbook['GIAO-VIEN']
            
            # Clear existing data (keep headers)
            teachers_sheet.delete_rows(2, teachers_sheet.max_row)
            
            # Điền dữ liệu từ row 2
            for idx, row in self.teachers_df.iterrows():
                row_num = idx + 2
                
                teachers_sheet[f'A{row_num}'] = row['STT']
                teachers_sheet[f'B{row_num}'] = row['Tên giáo viên']
                teachers_sheet[f'C{row_num}'] = row['Ngày sinh']
                teachers_sheet[f'D{row_num}'] = row['Tên đăng nhập']
                teachers_sheet[f'E{row_num}'] = row['Mật khẩu đăng nhập lần đầu']
            
            # Áp dụng border và alignment chuẩn như base_processor
            max_data_row = len(self.teachers_df) + 1  # +1 cho header
            max_data_col = 5  # 5 cột: STT, Tên, Ngày sinh, Tên đăng nhập, Mật khẩu
            
            # Các cột cần căn giữa: STT (1), Ngày sinh (3), Mật khẩu (5)
            center_columns = [1, 3, 5]
            self.apply_border_to_sheet(teachers_sheet, max_data_row, max_data_col, center_columns)
            
            # Auto-adjust column widths theo chuẩn Mode 1
            column_widths = {
                'A': 8,   # STT
                'B': 30,  # Tên giáo viên - rộng hơn cho tên dài
                'C': 15,  # Ngày sinh
                'D': 40,  # Tên đăng nhập - rộng hơn cho email
                'E': 30   # Mật khẩu
            }
            for col, width in column_widths.items():
                teachers_sheet.column_dimensions[col].width = width
                
            # Set row height để text hiển thị đẹp
            for row_num in range(1, len(self.teachers_df) + 2):
                teachers_sheet.row_dimensions[row_num].height = 20
            
            # Căn giữa toàn bộ header (hàng 1)
            for col in range(1, max_data_col + 1):
                cell = teachers_sheet.cell(row=1, column=col)
                cell.alignment = self.center_alignment

            print(f"✅ Đã điền {len(self.teachers_df)} giáo viên vào sheet GIAO-VIEN")
            return True
            
        except Exception as e:
            print(f"❌ Lỗi khi điền dữ liệu giáo viên: {e}")
            return False
    
    def fill_students_sheet(self, workbook):
        """Điền dữ liệu học sinh vào sheet HOC-SINH"""
        try:
            if self.students_df is None or self.students_df.empty:
                print("⚠️ Không có dữ liệu học sinh để điền")
                return True
                
            students_sheet = workbook['HOC-SINH']
            
            # Clear existing data (keep headers)
            students_sheet.delete_rows(2, students_sheet.max_row)
            
            # Điền dữ liệu từ row 2
            for idx, row in self.students_df.iterrows():
                row_num = idx + 2
                
                students_sheet[f'A{row_num}'] = row['STT']
                # students_sheet[f'B{row_num}'] = row['Mã học sinh']
                students_sheet[f'C{row_num}'] = row['Họ và tên']
                students_sheet[f'D{row_num}'] = row['Ngày sinh']
                students_sheet[f'E{row_num}'] = row['Khối']
                students_sheet[f'F{row_num}'] = row['Lớp']
                students_sheet[f'G{row_num}'] = row['Tài khoản']
                students_sheet[f'H{row_num}'] = row['Mật khẩu lần đầu']
                students_sheet[f'I{row_num}'] = row['Mã đăng nhập cho PH']
            
            # Áp dụng border và alignment chuẩn như base_processor
            max_data_row = len(self.students_df) + 1  # +1 cho header
            max_data_col = 9
            
            # Các cột cần căn giữa: STT (1), Ngày sinh (3), Lớp (4), Mật khẩu (6), Mã PH (7)
            center_columns = [1, 2, 4, 5, 6, 8, 9]
            self.apply_border_to_sheet(students_sheet, max_data_row, max_data_col, center_columns)
            
            # Auto-adjust column widths theo chuẩn Mode 1
            column_widths = {
                'A': 8,    # STT
                # 'B': 0,   # Mã học sinh
                'C': 30,   # Họ và tên
                'D': 15,   # Ngày sinh
                'E': 10,   # Khối
                'F': 15,   # Lớp
                'G': 30,   # Tài khoản
                'H': 20,   # Mật khẩu lần đầu
                'I': 25    # Mã đăng nhập cho PH
            }
            for col, width in column_widths.items():
                students_sheet.column_dimensions[col].width = width
                
            # Set row height để text hiển thị đẹp
            for row_num in range(1, len(self.students_df) + 2):
                students_sheet.row_dimensions[row_num].height = 20
            
            # Căn giữa toàn bộ header (hàng 1)
            for col in range(1, max_data_col + 1):
                cell = students_sheet.cell(row=1, column=col)
                cell.alignment = self.center_alignment

            print(f"✅ Đã điền {len(self.students_df)} học sinh vào sheet HOC-SINH")
            return True
            
        except Exception as e:
            print(f"❌ Lỗi khi điền dữ liệu học sinh: {e}")
            return False
    
    def create_excel_output(self, output_path: str = None):
        """Tạo file Excel output theo template chuẩn Mode 1"""
        try:
            if output_path is None:
                # Tạo tên file với format "Export_Tên trường.xlsx"
                safe_school_name = "".join(c for c in self.school_name if c.isalnum() or c in (' ', '-')).strip()
                output_path = f"data/output/Export_{safe_school_name}.xlsx"
            
            # Đảm bảo thư mục output tồn tại
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Copy template làm base
            if not self.copy_template(output_path):
                return None
            
            # Load workbook
            workbook = load_workbook(output_path)
            
            # Cập nhật các sheet
            if not self.update_admin_sheet(workbook):
                return None
                
            if not self.fill_teachers_sheet(workbook):
                return None
                
            if not self.fill_students_sheet(workbook):
                return None
            
            # Save workbook
            workbook.save(output_path)
            workbook.close()
            
            print(f"🎉 Đã tạo thành công file Excel: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"❌ Lỗi khi tạo file Excel: {e}")
            return None
    
    def convert(self, output_path: str = None):
        """Thực hiện chuyển đổi toàn bộ"""
        print("🚀 BẮT ĐẦU CHUYỂN ĐỔI JSON SANG EXCEL (TEMPLATE MODE 1)")
        print("=" * 60)
        
        # Load JSON data
        if not self.load_json_data():
            return None
        
        # Extract data
        if not self.extract_teachers_data():
            return None
            
        if not self.extract_students_data():
            return None
        
        # Create Excel output
        result_path = self.create_excel_output(output_path)
        
        if result_path:
            print("\n" + "=" * 60)
            print("🎊 CHUYỂN ĐỔI HOÀN TẤT!")
            print(f"📄 File Excel đầu ra: {result_path}")
            print(f"👥 Số giáo viên: {len(self.teachers_df) if self.teachers_df is not None else 0}")
            print(f"🎓 Số học sinh: {len(self.students_df) if self.students_df is not None else 0}")
            print(f"🏫 Trường: {self.school_name}")
        
        return result_path


def convert_json_to_excel_template(json_file_path: str, output_path: str = None, template_path: str = None):
    """
    Hàm tiện ích để chuyển đổi JSON sang Excel template format
    
    Args:
        json_file_path (str): Đường dẫn file JSON
        output_path (str): Đường dẫn file Excel output (optional)
        template_path (str): Đường dẫn template Excel (optional)
    
    Returns:
        str: Đường dẫn file Excel output nếu thành công, None nếu lỗi
    """
    converter = JSONToExcelTemplateConverter(json_file_path, template_path)
    return converter.convert(output_path)


if __name__ == "__main__":
    # Test với file JSON workflow
    print("🧪 TEST JSON TO EXCEL TEMPLATE CONVERTER")
    print("=" * 50)
    
    # Đường dẫn file JSON (sử dụng file mới nhất)
    json_files = [
        "data/output/workflow_filtered_GDNN - GDTX TP Chí Linh_20250729_085609.json"
    ]
    
    for json_file in json_files:
        if os.path.exists(json_file):
            print(f"\n🔄 Đang chuyển đổi: {json_file}")
            result = convert_json_to_excel_template(json_file)
            
            if result:
                print(f"✅ Thành công: {result}")
            else:
                print("❌ Chuyển đổi thất bại")
        else:
            print(f"❌ File không tồn tại: {json_file}")
