#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
File: ui_helpers.py
Purpose: Helper methods cho UI main window
Author: School Process Application
Date: 2025-08-07
Location: ui/ui_helpers.py
"""

import pandas as pd
import os
import subprocess
import platform
from tkinter import messagebox
from openpyxl.styles import Alignment, Border, Side

def export_teachers_to_excel(teachers_list, school_name, school_year, log_callback=None):
    """Export danh sách giáo viên ra file Excel theo template"""
    try:
        def log(msg, level="info"):
            if log_callback:
                log_callback(msg, level)
            else:
                print(f"[{level}] {msg}")
        
        # Tạo dữ liệu theo cấu trúc template
        df_data = []
        for i, teacher in enumerate(teachers_list, 1):
            # Trích xuất dữ liệu từ cấu trúc API
            teacher_info = teacher.get('teacherInfo', {})
            
            row = {
                'STT': i,
                'Tên giáo viên': teacher_info.get('displayName', ''),
                'Ngày sinh': teacher_info.get('userBirthday', ''),
                'Tên đăng nhập': teacher_info.get('userName', ''),
                'Mật khẩu đăng nhập lần đầu': teacher_info.get('pwd', '')
            }
            df_data.append(row)
        
        # Tạo DataFrame
        df = pd.DataFrame(df_data)
        
        # Tạo tên file
        safe_school_name = "".join(c for c in school_name if c.isalnum() or c in (' ', '-', '_')).strip()
        filename = f"DSGV_{safe_school_name}-{school_year}.xlsx"
        output_path = f"data/output/{filename}"
        
        # Tạo thư mục nếu chưa có
        os.makedirs("data/output", exist_ok=True)
        
        # Lưu file Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='GIAO-VIEN', index=False)
            
            # Format Excel
            worksheet = writer.sheets['GIAO-VIEN']
            
            # Tạo border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Căn giữa header (dòng 1)
            for cell in worksheet[1]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Căn giữa các cột được chỉ định và thêm border
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                for cell in row:
                    # Thêm border cho tất cả cells
                    cell.border = thin_border
                    
                    # Căn giữa các cột cụ thể
                    col_name = worksheet.cell(row=1, column=cell.column).value
                    if col_name in ['STT', 'Ngày sinh', 'Mật khẩu đăng nhập lần đầu']:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        log(f"✅ Đã xuất {len(teachers_list)} giáo viên ra file Excel", "success")
        return output_path
        
    except Exception as e:
        log(f"❌ Lỗi export Excel: {str(e)}", "error")
        import traceback
        traceback.print_exc()
        return None


def export_students_to_excel(students_list, school_name, school_year, log_callback=None):
    """Export danh sách học sinh ra file Excel theo template"""
    try:
        def log(msg, level="info"):
            if log_callback:
                log_callback(msg, level)
            else:
                print(f"[{level}] {msg}")
        
        # Tạo dữ liệu theo cấu trúc template
        df_data = []
        for i, student in enumerate(students_list, 1):
            # Trích xuất dữ liệu từ cấu trúc API
            user_info = student.get('userInfo', {})
            
            # Lấy thông tin lớp từ groupClass
            group_class = student.get('groupClass', [])
            class_name = ''
            if group_class and len(group_class) > 0:
                class_name = group_class[0].get('className', '')
            
            row = {
                'STT': i,
                'Mã học sinh': student.get('studentId', ''),
                'Họ và tên': user_info.get('displayName', ''),
                'Ngày sinh': user_info.get('userBirthday', ''),
                'Khối': student.get('grade', ''),
                'Lớp': class_name,
                'Tài khoản': user_info.get('userName', ''),
                'Mật khẩu lần đầu': user_info.get('pwd', ''),
                'Mã đăng nhập cho PH': user_info.get('codePin', '')
            }
            df_data.append(row)
        
        # Tạo DataFrame
        df = pd.DataFrame(df_data)
        
        # Tạo tên file
        safe_school_name = "".join(c for c in school_name if c.isalnum() or c in (' ', '-', '_')).strip()
        filename = f"DSHS_{safe_school_name}-{school_year}.xlsx"
        output_path = f"data/output/{filename}"
        
        # Tạo thư mục nếu chưa có
        os.makedirs("data/output", exist_ok=True)
        
        # Lưu file Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Danh sách HS toàn trường', index=False)
            
            # Format Excel
            worksheet = writer.sheets['Danh sách HS toàn trường']
            
            # Tạo border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Căn giữa header (dòng 1)
            for cell in worksheet[1]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Căn giữa các cột được chỉ định và thêm border
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                for cell in row:
                    # Thêm border cho tất cả cells
                    cell.border = thin_border
                    
                    # Căn giữa các cột cụ thể
                    col_name = worksheet.cell(row=1, column=cell.column).value
                    if col_name in ['STT', 'Mã học sinh', 'Ngày sinh', 'Khối', 'Lớp', 'Mật khẩu lần đầu', 'Mã đăng nhập cho PH']:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        log(f"✅ Đã xuất {len(students_list)} học sinh ra file Excel", "success")
        return output_path
        
    except Exception as e:
        log(f"❌ Lỗi export Excel: {str(e)}", "error")
        import traceback
        traceback.print_exc()
        return None


def show_file_completion_dialog(file_path, data_type, log_callback=None):
    """Hiển thị dialog hoàn thành và hỏi có muốn mở file"""
    try:
        def log(msg, level="info"):
            if log_callback:
                log_callback(msg, level)
            else:
                print(f"[{level}] {msg}")
        
        result = messagebox.askyesno(
            "Hoàn thành", 
            f"✅ Đã tạo thành công file {data_type}!\n\n"
            f"📄 File: {os.path.basename(file_path)}\n"
            f"📁 Thư mục: {os.path.dirname(file_path)}\n\n"
            f"Bạn có muốn mở file Excel ngay bây giờ?"
        )
        
        if result:
            try:
                if platform.system() == 'Windows':
                    os.startfile(file_path)
                elif platform.system() == 'Darwin':  # macOS
                    subprocess.run(['open', file_path])
                else:  # Linux
                    subprocess.run(['xdg-open', file_path])
                
                log(f"✅ Đã mở file {data_type}", "success")
            except Exception as e:
                log(f"❌ Không thể mở file: {str(e)}", "error")
                messagebox.showerror("Lỗi", f"Không thể mở file:\n{str(e)}")
        
    except Exception as e:
        if log_callback:
            log_callback(f"❌ Lỗi hiển thị dialog: {str(e)}", "error")
        else:
            print(f"❌ Lỗi hiển thị dialog: {str(e)}")
